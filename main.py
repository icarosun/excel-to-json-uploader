import os
import json 
import openpyxl
import datetime
import requests

def excel_to_json(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    print(wb['15630000(4787)']['BY277'].value)

    # for sheet in wb.sheetnames:
    #     print(sheet.split("(")[0], wb[sheet]['BY278'].value)

def main():
    print("Os valores coletados aqui terão que ser padrão para todas os ficheiros.")

    #TODO: Use lib date to change the value for date

    # first_date = input("Digite a data da primeira previsão coletada (ex.: dd/mm/yyyy): ")
    first_date = "10/08/2025"
    # last_date = input("Digite a data da última previsão coletada (ex.: dd/mm/yyyy): ")
    last_date = "31/10/2025"
    # column_date = input("Digite a COLUNA da PLANILHA em que estão os respectivos valores da data (ex.: BP): ")
    column_date = "BP"

    # first_line = int(input("Digite o número da primeira LINHA DA PLANILHA da previsão coletada e normalizada (ex.: 277): "))
    first_line = 277
    # last_line = int(input("Digite o número da última LINHA DA PLANILHA da previsão coleteada e normalizada (ex.: 359): "))
    last_line = 359
    # column = input("Digite a COLUNA DA PLANILHA onde estão os respectivos valores de previsão coletada e normalizada (ex.: BY): ")
    column_forecast = 'BY'

    #TODO: Check the values in one sheet
    #TODO: Check if the date values in backend is the next
    #TODO: Check the values before go a head

    #excel_to_json("madeira.xlsx")
    wb = openpyxl.load_workbook("madeira.xlsx", data_only=True)

    #TODO: Get station

    print(wb['15630000(4787)']['BY278'].value)
    print(wb['15630000(4787)']['BP278'].value)

    station = '15630000'
    
    #TODO: Create a class File with the values collected before

    data_collected = {
        'station_id': station,
        'first_date': first_date,
        'last_date' : last_date,
        'forecasts': []
    }

    #TODO: Date is very boring, I have to check this sequence
    #TODO: First make a check for every sheet before create the object


    for line in range(first_line, last_line + 1):
        coordinates_date = f"{column_date}{line}"
        coordinates_forecast = f"{column_forecast}{line}"

        date = wb['15630000(4787)'][coordinates_date].value 
        forecast = wb['15630000(4787)'][coordinates_forecast].value

        if (date != None or forecast != None):
            object_forecast = {
                'date': date.strftime("%Y-%m-%d %H:%M:%S"),
                'forecast': forecast
            }

            data_collected['forecasts'].append(object_forecast)
        else:
            line = last_line + 1
            print("Stop")

    wb.close()

    url = f'http://0.0.0.0:3333/{station}'

    r = requests.post(url, data=json.dumps(data_collected))

    print(r.status_code)

if __name__ == "__main__":
    main()
